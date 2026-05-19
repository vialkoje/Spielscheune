#!/usr/bin/env python3
"""
Erstellt spielscheune.xlsx mit Kundenkartei und Monatsabrechnung.
Einmalig ausführen – danach in Excel als .xlsm speichern und Makro einrichten.
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR    = Path(__file__).parent
EXCEL_DATEI = BASE_DIR / "spielscheune.xlsx"

# ── Farben ─────────────────────────────────────────────────────────────────────
BLAU       = "1F4E79"
WEISS      = "FFFFFF"
GRAU_HELL  = "F2F2F2"
BLAU_HELL  = "D6E4F0"
GRUEN      = "375623"
GRUEN_HELL = "E2EFDA"

RAND_DÜNN = Border(bottom=Side(style="thin", color="DDDDDD"))


def _set(cell, **kwargs):
    for k, v in kwargs.items():
        setattr(cell, k, v)


def kopf_zeile(ws, headers: list[str], breiten: list[int]) -> None:
    ws.row_dimensions[1].height = 32
    for col, (h, b) in enumerate(zip(headers, breiten), 1):
        ws.column_dimensions[get_column_letter(col)].width = b
        c = ws.cell(row=1, column=col, value=h)
        _set(c,
             font=Font(bold=True, color=WEISS, size=11),
             fill=PatternFill("solid", fgColor=BLAU),
             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
             border=Border(bottom=Side(style="medium", color=WEISS)))


def daten_zeile(ws, zeile: int, n_cols: int) -> None:
    farbe = WEISS if zeile % 2 == 0 else GRAU_HELL
    ws.row_dimensions[zeile].height = 20
    for col in range(1, n_cols + 1):
        c = ws.cell(row=zeile, column=col)
        _set(c,
             fill=PatternFill("solid", fgColor=farbe),
             alignment=Alignment(vertical="center"),
             border=RAND_DÜNN)


# ── Blatt 1: Kundenkartei ──────────────────────────────────────────────────────

def erstelle_kundenkartei(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Kundenkartei")

    headers = ["Kundennummer", "Name", "Strasse", "PLZ", "Ort", "Rechnungsemail", "Abrechnungsfrequenz"]
    breiten  = [15, 26, 26, 7,  20, 32, 20]
    kopf_zeile(ws, headers, breiten)

    kunden = [
        ["K123", "Max Muster",     "Musterweg 1",     "12345", "Ort",           "jens@vialkowitsch.de", "monatlich"],
        ["K124", "Lisa Beispiel",  "Beispielgasse 5", "54321", "Stadt",         "jens@vialkowitsch.de", "vierteljährlich"],
        ["K125", "Tom Testeltern", "Teststraße 7",    "11223", "Beispielstadt", "jens@vialkowitsch.de", "monatlich"],
    ]

    for r, row in enumerate(kunden, 2):
        daten_zeile(ws, r, len(headers))
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)

    ws.freeze_panes = "A2"


# ── Blatt 2: Monatsabrechnung ──────────────────────────────────────────────────

XLOOKUP_FORMEL = (
    '=IF(A{r}="","",IFERROR('
    'XLOOKUP(A{r},Kundenkartei!$A:$A,Kundenkartei!$B:$B,"?"),"?"))'
)

def erstelle_monatsabrechnung(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Monatsabrechnung")

    headers = ["Kundennummer", "Name", "Kind_Vorname", "LZR", "Anzahl_Essen", "Preis_pro_Essen"]
    breiten  = [15, 26, 14, 14, 14, 16]
    kopf_zeile(ws, headers, breiten)

    eintraege = [
        ["K123", "Elias",  "April 2025", 20, 3.5],
        ["K124", "Jonas",  "April 2025", 60, 3.5],
        ["K125", "Jürgen", "April 2025", 18, 3.75],
        ["K125", "Emil",   "April 2025", 10, 3.5],
    ]

    for r, row in enumerate(eintraege, 2):
        daten_zeile(ws, r, len(headers))

        # A: Kundennummer (manuell)
        ws.cell(row=r, column=1, value=row[0])

        # B: Name via XLOOKUP – hellblau, schreibgeschützt für den Nutzer
        c_name = ws.cell(row=r, column=2, value=XLOOKUP_FORMEL.format(r=r))
        _set(c_name,
             fill=PatternFill("solid", fgColor=BLAU_HELL),
             font=Font(italic=True, color=BLAU, size=10),
             alignment=Alignment(vertical="center"),
             border=RAND_DÜNN)

        # C–F: Bewegungsdaten (manuell)
        for col, val in enumerate(row[1:], 3):
            c = ws.cell(row=r, column=col, value=val)
        ws.cell(row=r, column=5).number_format = "0"
        ws.cell(row=r, column=6).number_format = '#,##0.00 "€"'

    # XLOOKUP-Formel für Folgezeilen vorbereiten (bis Zeile 100)
    for r in range(len(eintraege) + 2, 101):
        c = ws.cell(row=r, column=2, value=XLOOKUP_FORMEL.format(r=r))
        _set(c,
             fill=PatternFill("solid", fgColor=BLAU_HELL),
             font=Font(italic=True, color=BLAU, size=10),
             alignment=Alignment(vertical="center"))

    ws.freeze_panes = "A2"

    # ── Button-Bereich (H2:J5) – bewusst UNTER Zeile 1, damit H1 leer bleibt ──
    # (openpyxl liest Zeile 1 als Spaltenheader; ein Wert in H1 würde als
    #  Spaltenname im DataFrame landen und den Datenimport stören)
    for col_letter in ("G", "H", "I", "J"):
        ws.column_dimensions[col_letter].width = 10

    # Grüner Button (H2:J4)
    ws.merge_cells("H2:J4")
    btn = ws["H2"]
    _set(btn,
         value="▶  Abrechnung\n    starten",
         font=Font(bold=True, color=WEISS, size=13),
         fill=PatternFill("solid", fgColor=GRUEN),
         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 14

    # Hinweiszeile (H5:J5)
    ws.merge_cells("H5:J5")
    hint = ws["H5"]
    _set(hint,
         value="⚙ Rechtsklick → Makro: StartAbrechnung",
         font=Font(italic=True, size=8, color="666666"),
         fill=PatternFill("solid", fgColor=GRUEN_HELL),
         alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[5].height = 14


# ── Hauptprogramm ──────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    erstelle_kundenkartei(wb)
    erstelle_monatsabrechnung(wb)

    wb.save(EXCEL_DATEI)
    print(f"✓ Erstellt: {EXCEL_DATEI}")
    print()
    print("Einmalige Einrichtung des Buttons:")
    print("  1. Öffne spielscheune.xlsx in Excel")
    print("  2. Datei → Speichern unter → Excel-Arbeitsmappe mit Makros (.xlsm)")
    print("     Dateiname: spielscheune.xlsm")
    print("  3. Extras → Makros → Visual Basic Editor (oder Alt+F11)")
    print("  4. Datei → Importieren → AbrechnungMakro.bas")
    print("  5. VBA-Editor schließen")
    print("  6. Rechtsklick auf grünen Button → Makro zuweisen → StartAbrechnung → OK")


if __name__ == "__main__":
    main()
